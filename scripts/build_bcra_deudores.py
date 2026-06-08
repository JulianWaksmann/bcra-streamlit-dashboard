#!/usr/bin/env python3
"""Build dashboard-ready outputs from BCRA Central de Deudores files.

The source files are fixed-width, encoded as latin-1-ish bytes, and amounts are
reported in thousands of pesos with one decimal using a comma separator.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import defaultdict
from pathlib import Path
from typing import Iterable


DEUDORES_LAYOUT = [
    ("codigo_entidad", 5),
    ("fecha_info", 6),
    ("tipo_id", 2),
    ("nro_id", 11),
    ("actividad_codigo", 3),
    ("situacion_codigo", 2),
    ("prestamos_total_garantias_afrontadas", 12),
    ("sin_uso", 12),
    ("garantias_otorgadas", 12),
    ("otros_conceptos", 12),
    ("garantias_preferidas_a", 12),
    ("garantias_preferidas_b", 12),
    ("sin_garantias_preferidas", 12),
    ("contragarantias_preferidas_a", 12),
    ("contragarantias_preferidas_b", 12),
    ("sin_contragarantias_preferidas", 12),
    ("previsiones", 12),
    ("deuda_cubierta", 1),
    ("proceso_judicial_revision", 1),
    ("refinanciaciones", 1),
    ("recategorizacion_obligatoria", 1),
    ("situacion_juridica", 1),
    ("irrecuperable_ult_parrafo", 1),
    ("dias_atraso", 4),
]

EXPECTED_DEUDORES_LEN = sum(width for _, width in DEUDORES_LAYOUT)

SITUACION_LABELS = {
    "1": "Situacion normal",
    "2": "Con seguimiento especial / riesgo bajo",
    "3": "Con problemas / riesgo medio",
    "4": "Con alto riesgo de insolvencia / riesgo alto",
    "5": "Irrecuperable",
    "11": "Asistencias totalmente cubiertas con garantias preferidas A",
}

AMOUNT_FIELDS = [
    "prestamos_total_garantias_afrontadas",
    "sin_uso",
    "garantias_otorgadas",
    "otros_conceptos",
    "garantias_preferidas_a",
    "garantias_preferidas_b",
    "sin_garantias_preferidas",
    "contragarantias_preferidas_a",
    "contragarantias_preferidas_b",
    "sin_contragarantias_preferidas",
    "previsiones",
]

SLICES = {}
_pos = 0
for _name, _width in DEUDORES_LAYOUT:
    SLICES[_name] = slice(_pos, _pos + _width)
    _pos += _width


def read_fixed_lines(path: Path) -> Iterable[str]:
    with path.open("rb") as handle:
        for raw in handle:
            yield raw.decode("latin-1").rstrip("\r\n")


def clean_text(value: str) -> str:
    return " ".join(value.strip().split())


def parse_int(value: str) -> int:
    value = value.strip()
    return int(value) if value else 0


def parse_amount_tenths(raw: str) -> int:
    """Return amount as tenths of thousands of pesos.

    Examples:
    - "118,0       " -> 1180
    - ",0          " -> 0
    - "0           " -> 0
    """

    value = raw.strip()
    if not value:
        return 0

    sign = -1 if value.startswith("-") else 1
    if sign == -1:
        value = value[1:].strip()

    if "," in value:
        whole, frac = value.split(",", 1)
    else:
        whole, frac = value, "0"

    whole = whole.strip() or "0"
    frac = (frac.strip() + "0")[0]
    return sign * (int(whole) * 10 + int(frac))


def amount_to_csv(value_tenths: int) -> str:
    sign = "-" if value_tenths < 0 else ""
    value_tenths = abs(value_tenths)
    whole, frac = divmod(value_tenths, 10)
    return f"{sign}{whole}.{frac}"


def period_date(period: str) -> str:
    if len(period) == 6 and period.isdigit():
        return f"{period[:4]}-{period[4:]}-01"
    return ""


def parse_deudor_line(line: str) -> dict[str, str | int]:
    row: dict[str, str | int] = {}
    for name, _width in DEUDORES_LAYOUT:
        raw = line[SLICES[name]]
        if name in AMOUNT_FIELDS:
            row[f"{name}_miles_pesos"] = amount_to_csv(parse_amount_tenths(raw))
        elif name == "situacion_codigo":
            row[name] = str(parse_int(raw))
        elif name in {
            "deuda_cubierta",
            "proceso_judicial_revision",
            "refinanciaciones",
            "recategorizacion_obligatoria",
            "situacion_juridica",
            "irrecuperable_ult_parrafo",
            "dias_atraso",
        }:
            row[name] = parse_int(raw)
        else:
            row[name] = raw.strip()

    row["fecha_info_date"] = period_date(str(row["fecha_info"]))
    row["exposicion_total_miles_pesos"] = amount_to_csv(
        parse_amount_tenths(line[SLICES["prestamos_total_garantias_afrontadas"]])
        + parse_amount_tenths(line[SLICES["garantias_otorgadas"]])
        + parse_amount_tenths(line[SLICES["otros_conceptos"]])
    )
    return row


def write_dim_entidades(input_dir: Path, output_dir: Path) -> Path:
    source = input_dir / "Maeent.txt"
    target = output_dir / "dim_entidades.csv"
    with target.open("w", encoding="utf-8", newline="") as out:
        writer = csv.writer(out)
        writer.writerow(["codigo_entidad", "nombre_entidad"])
        for line in read_fixed_lines(source):
            writer.writerow([line[:5].strip(), clean_text(line[5:])])
    return target


def write_dim_deudores(input_dir: Path, output_dir: Path) -> Path:
    source = input_dir / "Nomdeu.txt"
    target = output_dir / "dim_deudores_nomdeu.csv"
    with target.open("w", encoding="utf-8", newline="") as out:
        writer = csv.writer(out)
        writer.writerow(["nro_id", "denominacion"])
        for line in read_fixed_lines(source):
            writer.writerow([line[:11].strip(), clean_text(line[11:])])
    return target


def write_dim_situacion(output_dir: Path) -> Path:
    target = output_dir / "dim_situacion.csv"
    with target.open("w", encoding="utf-8", newline="") as out:
        writer = csv.writer(out)
        writer.writerow(["situacion_codigo", "situacion_descripcion"])
        for code, label in SITUACION_LABELS.items():
            writer.writerow([code, label])
    return target


def new_accumulator() -> list[int]:
    return [0] * 15


def update_accumulator(acc: list[int], line: str) -> None:
    prestamos = parse_amount_tenths(line[SLICES["prestamos_total_garantias_afrontadas"]])
    garantias = parse_amount_tenths(line[SLICES["garantias_otorgadas"]])
    otros = parse_amount_tenths(line[SLICES["otros_conceptos"]])
    previsiones = parse_amount_tenths(line[SLICES["previsiones"]])
    proceso = line[SLICES["proceso_judicial_revision"]].strip()
    dias = parse_int(line[SLICES["dias_atraso"]])

    acc[0] += 1
    acc[1] += prestamos
    acc[2] += garantias
    acc[3] += otros
    acc[4] += prestamos + garantias + otros
    acc[5] += previsiones
    acc[6] += 1 if line[SLICES["deuda_cubierta"]].strip() == "1" else 0
    acc[7] += 1 if proceso == "1" else 0
    acc[8] += 1 if proceso == "2" else 0
    acc[9] += 1 if line[SLICES["refinanciaciones"]].strip() == "1" else 0
    acc[10] += 1 if line[SLICES["recategorizacion_obligatoria"]].strip() == "1" else 0
    acc[11] += 1 if line[SLICES["situacion_juridica"]].strip() == "1" else 0
    acc[12] += 1 if line[SLICES["irrecuperable_ult_parrafo"]].strip() == "1" else 0
    acc[13] += dias
    acc[14] = max(acc[14], dias)


AGG_METRIC_HEADERS = [
    "registros",
    "prestamos_total_garantias_afrontadas_miles_pesos",
    "garantias_otorgadas_miles_pesos",
    "otros_conceptos_miles_pesos",
    "exposicion_total_miles_pesos",
    "previsiones_miles_pesos",
    "registros_deuda_cubierta",
    "registros_proceso_judicial",
    "registros_en_revision",
    "registros_refinanciados",
    "registros_recategorizacion_obligatoria",
    "registros_situacion_juridica",
    "registros_irrecuperable_ult_parrafo",
    "dias_atraso_promedio",
    "dias_atraso_max",
]


def accumulator_values(acc: list[int]) -> list[str | int]:
    avg_dias = "0.0" if acc[0] == 0 else f"{acc[13] / acc[0]:.1f}"
    return [
        acc[0],
        amount_to_csv(acc[1]),
        amount_to_csv(acc[2]),
        amount_to_csv(acc[3]),
        amount_to_csv(acc[4]),
        amount_to_csv(acc[5]),
        acc[6],
        acc[7],
        acc[8],
        acc[9],
        acc[10],
        acc[11],
        acc[12],
        avg_dias,
        acc[14],
    ]


def write_aggregate(path: Path, key_headers: list[str], data: dict[tuple[str, ...], list[int]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as out:
        writer = csv.writer(out)
        writer.writerow(key_headers + AGG_METRIC_HEADERS)
        for key in sorted(data):
            writer.writerow(list(key) + accumulator_values(data[key]))


def write_sample(output_dir: Path, rows: list[dict[str, str | int]]) -> Path:
    target = output_dir / "fact_deudas_sample.csv"
    if not rows:
        return target
    headers = list(rows[0].keys())
    with target.open("w", encoding="utf-8", newline="") as out:
        writer = csv.DictWriter(out, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
    return target


def scan_deudores(input_dir: Path, output_dir: Path, sample_rows: int, progress_every: int) -> dict[str, Path | int]:
    source = input_dir / "deudores.txt"
    bad_lengths = 0
    total = 0
    sample: list[dict[str, str | int]] = []

    by_situacion: dict[tuple[str, str], list[int]] = defaultdict(new_accumulator)
    by_entidad: dict[tuple[str, str], list[int]] = defaultdict(new_accumulator)
    by_entidad_situacion: dict[tuple[str, str, str], list[int]] = defaultdict(new_accumulator)
    by_actividad_situacion: dict[tuple[str, str, str], list[int]] = defaultdict(new_accumulator)
    by_entidad_actividad_situacion: dict[tuple[str, str, str, str], list[int]] = defaultdict(new_accumulator)

    for line in read_fixed_lines(source):
        total += 1
        if len(line) != EXPECTED_DEUDORES_LEN:
            bad_lengths += 1
            continue

        if len(sample) < sample_rows:
            sample.append(parse_deudor_line(line))

        periodo = line[SLICES["fecha_info"]].strip()
        entidad = line[SLICES["codigo_entidad"]].strip()
        actividad = line[SLICES["actividad_codigo"]].strip() or "SIN_DATO"
        situacion = str(parse_int(line[SLICES["situacion_codigo"]]))

        update_accumulator(by_situacion[(periodo, situacion)], line)
        update_accumulator(by_entidad[(periodo, entidad)], line)
        update_accumulator(by_entidad_situacion[(periodo, entidad, situacion)], line)
        update_accumulator(by_actividad_situacion[(periodo, actividad, situacion)], line)
        update_accumulator(by_entidad_actividad_situacion[(periodo, entidad, actividad, situacion)], line)

        if progress_every and total % progress_every == 0:
            print(f"processed {total:,} rows", file=sys.stderr)

    outputs: dict[str, Path | int] = {
        "total_rows": total,
        "bad_length_rows": bad_lengths,
        "sample": write_sample(output_dir, sample),
    }
    aggregate_specs = [
        ("agg_situacion.csv", ["fecha_info", "situacion_codigo"], by_situacion),
        ("agg_entidad.csv", ["fecha_info", "codigo_entidad"], by_entidad),
        (
            "agg_entidad_situacion.csv",
            ["fecha_info", "codigo_entidad", "situacion_codigo"],
            by_entidad_situacion,
        ),
        (
            "agg_actividad_situacion.csv",
            ["fecha_info", "actividad_codigo", "situacion_codigo"],
            by_actividad_situacion,
        ),
        (
            "agg_entidad_actividad_situacion.csv",
            ["fecha_info", "codigo_entidad", "actividad_codigo", "situacion_codigo"],
            by_entidad_actividad_situacion,
        ),
    ]

    for filename, headers, data in aggregate_specs:
        target = output_dir / filename
        write_aggregate(target, headers, data)
        outputs[filename] = target

    return outputs


def write_bigquery_schema(output_dir: Path) -> Path:
    schema = [
        {"name": "fecha_info", "type": "STRING", "mode": "REQUIRED"},
        {"name": "fecha_info_date", "type": "DATE", "mode": "NULLABLE"},
        {"name": "codigo_entidad", "type": "STRING", "mode": "REQUIRED"},
        {"name": "tipo_id", "type": "STRING", "mode": "NULLABLE"},
        {"name": "nro_id", "type": "STRING", "mode": "NULLABLE"},
        {"name": "actividad_codigo", "type": "STRING", "mode": "NULLABLE"},
        {"name": "situacion_codigo", "type": "INT64", "mode": "NULLABLE"},
    ]
    for field in AMOUNT_FIELDS:
        schema.append({"name": f"{field}_miles_pesos", "type": "NUMERIC", "mode": "NULLABLE"})
    schema.extend(
        [
            {"name": "exposicion_total_miles_pesos", "type": "NUMERIC", "mode": "NULLABLE"},
            {"name": "deuda_cubierta", "type": "INT64", "mode": "NULLABLE"},
            {"name": "proceso_judicial_revision", "type": "INT64", "mode": "NULLABLE"},
            {"name": "refinanciaciones", "type": "INT64", "mode": "NULLABLE"},
            {"name": "recategorizacion_obligatoria", "type": "INT64", "mode": "NULLABLE"},
            {"name": "situacion_juridica", "type": "INT64", "mode": "NULLABLE"},
            {"name": "irrecuperable_ult_parrafo", "type": "INT64", "mode": "NULLABLE"},
            {"name": "dias_atraso", "type": "INT64", "mode": "NULLABLE"},
        ]
    )
    target = output_dir / "bigquery_fact_deudas_schema.json"
    target.write_text(json.dumps(schema, indent=2) + "\n", encoding="utf-8")
    return target


def write_workbook(output_dir: Path) -> Path | None:
    try:
        from openpyxl import Workbook
    except ImportError:
        return None

    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    csv_files = [
        "agg_situacion.csv",
        "agg_entidad.csv",
        "agg_entidad_situacion.csv",
        "agg_actividad_situacion.csv",
        "agg_entidad_actividad_situacion.csv",
        "dim_entidades.csv",
        "dim_situacion.csv",
    ]
    for filename in csv_files:
        path = output_dir / filename
        if not path.exists():
            continue
        sheet = workbook.create_sheet(filename.removesuffix(".csv")[:31])
        with path.open("r", encoding="utf-8", newline="") as handle:
            for row in csv.reader(handle):
                sheet.append(row)

    notes = workbook.create_sheet("notas")
    notes.append(["clave", "valor"])
    notes.append(["fuente", "BCRA Central de Deudores del Sistema Financiero"])
    notes.append(["periodo", "202604"])
    notes.append(["unidad_montos", "miles de pesos con un decimal"])
    notes.append(["grain_detalle", "una fila por entidad, periodo, identificacion y situacion informada"])
    notes.append(["exposicion_total", "prestamos_total_garantias_afrontadas + garantias_otorgadas + otros_conceptos"])
    notes.append(["privacidad", "para publicacion se recomienda usar agregados, no identificaciones personales"])

    target = output_dir / "dashboard_bcra_deudores_202604.xlsx"
    workbook.save(target)
    return target


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", type=Path, default=Path("202604"))
    parser.add_argument("--output-dir", type=Path, default=Path("202604/processed"))
    parser.add_argument("--sample-rows", type=int, default=20000)
    parser.add_argument("--progress-every", type=int, default=5_000_000)
    args = parser.parse_args()

    args.output_dir.mkdir(parents=True, exist_ok=True)

    outputs: dict[str, Path | int | None] = {}
    outputs["dim_entidades"] = write_dim_entidades(args.input_dir, args.output_dir)
    outputs["dim_deudores_nomdeu"] = write_dim_deudores(args.input_dir, args.output_dir)
    outputs["dim_situacion"] = write_dim_situacion(args.output_dir)
    outputs.update(scan_deudores(args.input_dir, args.output_dir, args.sample_rows, args.progress_every))
    outputs["bigquery_schema"] = write_bigquery_schema(args.output_dir)
    outputs["workbook"] = write_workbook(args.output_dir)

    manifest = args.output_dir / "manifest.json"
    manifest.write_text(
        json.dumps({key: str(value) for key, value in outputs.items()}, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps({key: str(value) for key, value in outputs.items()}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
